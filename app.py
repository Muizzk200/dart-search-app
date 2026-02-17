from flask import Flask, render_template, request, jsonify, send_file, g
import io
import openpyxl
import os
import time
import tempfile
import uuid
import traceback
import logging
from werkzeug.utils import secure_filename

app = Flask(__name__)
# Use a temp dir for uploads on Render (safer for ephemeral containers)
app.config['UPLOAD_FOLDER'] = os.path.join(tempfile.gettempdir(), 'dart_uploads')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
ALLOWED_EXTENSIONS = {'xlsx', 'csv'}
BLANK_LABEL = '(blank)'

# Store loaded data in memory
loaded_data = {
    'rows': [],
    'filename': None
}

# Compact field order used for tuple storage. Keep order matching front-end expectations.
FIELDS = [
    'item_no',
    'description',
    'product_division',
    'material_group',
    'material_group_desc',
    'manufacturer_name',
    'manufacturer_item_no',
    'sales_status',
    'product_manager',
    'sub_item'
]
# Map field name -> index for fast access
FIELD_IDX = {name: i for i, name in enumerate(FIELDS)}

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Ensure the uploads directory exists at module import time so the directory
# is present when the app is started by a WSGI server (e.g. gunicorn on Render).
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Basic logger setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('dart_app')


# Request/response logging middleware
@app.before_request
def start_request():
    # assign a request id
    rid = str(uuid.uuid4())
    g.request_id = rid
    safe_headers = {}
    for h in ['Host', 'User-Agent', 'Content-Type', 'Accept', 'Content-Length']:
        v = request.headers.get(h)
        if v:
            safe_headers[h] = v
    logger.info(f"[RID:{rid}] Incoming {request.method} {request.path} headers={safe_headers}")


@app.after_request
def after_request(response):
    rid = getattr(g, 'request_id', 'N/A')
    try:
        status = response.status_code
        if status >= 400:
            # try to get response data (may be bytes/stream); limit size
            try:
                body = (response.get_data(as_text=True) or '')[:1000]
            except Exception:
                body = '<unavailable>'
            logger.warning(f"[RID:{rid}] Response {status} for {request.method} {request.path}: {body}")
        else:
            logger.info(f"[RID:{rid}] Response {status} for {request.method} {request.path}")
    except Exception:
        logger.exception(f"[RID:{rid}] after_request logging failed")
    return response


@app.errorhandler(Exception)
def handle_exception(e):
    rid = getattr(g, 'request_id', str(uuid.uuid4()))
    tb = traceback.format_exc()
    logger.error(f"[RID:{rid}] Unhandled exception during {request.method} {request.path}: {str(e)}\n{tb}")
    # Return JSON error with request id so client can report it
    return jsonify({'success': False, 'message': 'Internal server error', 'request_id': rid}), 500


def build_header_index(header_row):
    """
    Build a dynamic header → column index mapping from the first row.
    Only 'description' is mandatory. Missing columns are mapped to -1.
    
    Args:
        header_row: List of header values from row 1
        
    Returns:
        tuple: (header_index dict, error_message)
        - header_index: dict mapping normalized header names to column indices (-1 if missing)
        - error_message: None if successful, error string if description column is missing
        
    Example:
        header_index = {
            'item': 0,
            'description': 1,
            'product division': 5,
            'product mgr': -1,  # missing column
            ...
        }
    """
    REQUIRED_HEADERS = ['description']
    
    EXPECTED_HEADERS = [
        'item',
        'description',
        'product division',
        'sales status',
        'mfr name',
        'mfr item',
        'sub item',
        'product mgr',
        'material group',
        'material group desc'
    ]
    
    # Build normalized header map (lowercase, stripped)
    header_index = {}
    for idx, header in enumerate(header_row):
        if header is not None:
            normalized = str(header).strip().lower()
            header_index[normalized] = idx
    
    # Check if required columns exist (only description is mandatory)
    for required_header in REQUIRED_HEADERS:
        if required_header not in header_index:
            return None, f"Missing required column: {required_header}"
    
    # Add missing expected columns with -1 index to indicate they don't exist
    for expected_header in EXPECTED_HEADERS:
        if expected_header not in header_index:
            header_index[expected_header] = -1
    
    return header_index, None

def parse_excel_file(filepath):
    """
    Parse Excel file and extract data from DART sheet using header-based column mapping.
    
    Column headers are read from row 1 and normalized (stripped, lowercased).
    All data extraction uses the dynamic header_index mapping instead of fixed column positions.
    
    Expected headers (case-insensitive):
    - Item
    - Description
    - Product Division
    - Sales Status
    - Mfr Name
    - Mfr Item
    - Sub Item
    - Product Mgr
    """
    try:
        # Use read_only to reduce memory usage and speed up large files
        workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        # Try to find 'DART' sheet, fallback to first sheet
        if 'DART' in workbook.sheetnames:
            worksheet = workbook['DART']
        else:
            worksheet = workbook.active

        # Read header row (row 1)
        header_row = None
        for idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
            header_row = row
            break
        
        if not header_row:
            return [], "No header row found in Excel file"
        
        # Build header → index mapping and validate required headers exist
        header_index, error = build_header_index(header_row)
        if error:
            return [], error

        rows = []
        # Process data rows starting from row 2
        # We now read all columns since we use header-based indexing
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            try:
                # Skip empty rows (check if description is empty)
                desc_idx = header_index['description']
                short_desc = row[desc_idx] if len(row) > desc_idx else None
                if not short_desc or not str(short_desc).strip():
                    continue
                
                # Extract values using header-based index mapping
                # Handle missing columns gracefully by returning empty string for missing values
                def val_for(hdr_key):
                    idx = header_index.get(hdr_key)
                    if idx is None or idx == -1:
                        return ''
                    if len(row) <= idx:
                        return ''
                    v = row[idx]
                    return '' if v is None else v

                item_no = val_for('item')
                product_div = val_for('product division')
                material_group = val_for('material group')
                material_group_desc = val_for('material group desc')
                mfg_name = val_for('mfr name')
                mfg_item_no = val_for('mfr item')
                sales_status = val_for('sales status')
                product_manager = val_for('product mgr')
                sub_item = val_for('sub item')

                # Store as compact tuple in FIELDS order
                rows.append((
                    item_no,
                    short_desc,
                    product_div,
                    material_group,
                    material_group_desc,
                    mfg_name,
                    mfg_item_no,
                    sales_status,
                    product_manager,
                    sub_item
                ))
            except (IndexError, TypeError):
                continue

        workbook.close()
        return rows, None

    except Exception as e:
        return [], f"Error parsing file: {str(e)}"


def parse_csv_file(filepath):
    """
    Parse a CSV file using header-based column mapping.
    
    The first row is treated as headers and normalized (stripped, lowercased).
    All data extraction uses the dynamic header_index mapping.
    
    Expected headers (case-insensitive):
    - Item
    - Description
    - Product Division
    - Sales Status
    - Mfr Name
    - Mfr Item
    - Sub Item
    - Product Mgr
    """
    import csv

    rows = []
    try:
        with open(filepath, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            
            # Read header row (first row)
            header_row = next(reader, None)
            if not header_row:
                return [], "No header row found in CSV file"
            
            # Build header → index mapping and validate required headers exist
            header_index, error = build_header_index(header_row)
            if error:
                return [], error
            
            # Process data rows
            for r in reader:
                try:
                    # Skip empty rows (check if description is empty)
                    desc_idx = header_index['description']
                    short_desc = r[desc_idx] if len(r) > desc_idx else None
                    if not short_desc or not str(short_desc).strip():
                        continue
                    
                    # Extract values using header-based index mapping
                    def val_for(hdr_key):
                        idx = header_index.get(hdr_key)
                        if idx is None or idx == -1:
                            return ''
                        if len(r) <= idx:
                            return ''
                        v = r[idx]
                        return '' if v is None else v

                    item_no = val_for('item')
                    product_div = val_for('product division')
                    material_group = val_for('material group')
                    material_group_desc = val_for('material group desc')
                    mfg_name = val_for('mfr name')
                    mfg_item_no = val_for('mfr item')
                    sales_status = val_for('sales status')
                    product_manager = val_for('product mgr')
                    sub_item = val_for('sub item')

                    rows.append((
                        item_no,
                        short_desc,
                        product_div,
                        material_group,
                        material_group_desc,
                        mfg_name,
                        mfg_item_no,
                        sales_status,
                        product_manager,
                        sub_item
                    ))
                except Exception:
                    continue
        
        return rows, None
    except Exception as e:
        return [], f"Error parsing CSV file: {str(e)}"


def _compute_filters(rows):
    """Compute unique filter options from parsed rows."""
    manufacturers = set()
    product_divs = set()
    sales_statuses = set()
    product_managers = set()
    sub_items = set()
    material_groups = set()
    material_group_descs = set()

    # rows are stored as tuples in FIELDS order; use FIELD_IDX mapping
    for r in rows:
        def get_val(field):
            v = r[FIELD_IDX[field]] if len(r) > FIELD_IDX[field] else ''
            return v

        mn = get_val('manufacturer_name')
        if mn:
            manufacturers.add(str(mn).strip())
        pd = get_val('product_division')
        if pd:
            product_divs.add(str(pd).strip())
        ss = get_val('sales_status')
        if ss is not None and str(ss).strip() != '':
            sales_statuses.add(str(ss).strip())
        else:
            sales_statuses.add(BLANK_LABEL)
        pm = get_val('product_manager')
        if pm:
            product_managers.add(str(pm).strip())
        si = get_val('sub_item')
        if si:
            sub_items.add(str(si).strip())
        mg = get_val('material_group')
        if mg:
            material_groups.add(str(mg).strip())
        mgd = get_val('material_group_desc')
        if mgd:
            material_group_descs.add(str(mgd).strip())

    return {
        'manufacturers': sorted([m for m in manufacturers if m]),
        'product_divisions': sorted([p for p in product_divs if p]),
        'sales_statuses': sorted([s for s in sales_statuses if s]),
        'product_managers': sorted([pm for pm in product_managers if pm]),
        'sub_items': sorted([si for si in sub_items if si]),
        'material_groups': sorted([mg for mg in material_groups if mg]),
        'material_group_descs': sorted([mgd for mgd in material_group_descs if mgd])
    }

def search_rows(keywords, rows):
    """
    Search rows by Description and Item No fields (case-insensitive, partial matches allowed).
    
    Matching Rule:
    ALL keywords must be found in EITHER:
      - Description field, OR
      - Item No field
    
    Keywords cannot be split across fields. Each field is evaluated independently.
    
    Examples:
      Item: A12345-B, Description: "Steel Hex Bolt"
      Search: "A123"        → MATCH (found in Item)
      Search: "steel bolt"  → MATCH (found in Description)
      Search: "A123 bolt"   → NO MATCH (keywords split across fields)
    """
    if not keywords or not keywords.strip():
        return []
    
    # Split keywords and convert to lowercase
    search_words = keywords.lower().split()
    
    if not search_words:
        return []
    
    results = []

    desc_idx = FIELD_IDX['description']
    item_idx = FIELD_IDX['item_no']

    for row in rows:
        desc_text = ''
        item_text = ''
        try:
            desc_val = row[desc_idx] if len(row) > desc_idx else ''
            item_val = row[item_idx] if len(row) > item_idx else ''
            desc_text = str(desc_val).lower() if desc_val is not None else ''
            item_text = str(item_val).lower() if item_val is not None else ''
        except Exception:
            desc_text = ''
            item_text = ''

        searchable_texts = [desc_text, item_text]

        match_found = any(
            all(word in text for word in search_words)
            for text in searchable_texts
        )

        if match_found:
            results.append(row)

    return results

@app.route('/')
def index():
    """Render the main page"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'Only .xlsx files are allowed'}), 400
        
        # Save file
        filename = secure_filename(file.filename)
        # Ensure uploads folder exists (create just before saving)
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        t0 = time.time()
        file.save(filepath)
        t1 = time.time()
        print(f"Saved uploaded file to {filepath} in {t1 - t0:.2f}s")

        # Parse file depending on extension
        ext = filename.rsplit('.', 1)[1].lower()
        # Parse file depending on extension (timing logged)
        parse_start = time.time()
        if ext == 'csv':
            rows, error = parse_csv_file(filepath)
        else:
            rows, error = parse_excel_file(filepath)
        parse_end = time.time()
        if error is None:
            print(f"Parsed file {filename} in {parse_end - parse_start:.2f}s, rows={len(rows)}")
        
        if error:
            return jsonify({'success': False, 'message': error}), 400
        
        # Store in memory (rows are tuples) and compute filter options
        loaded_data['rows'] = rows
        loaded_data['filename'] = filename
        loaded_data['filters'] = _compute_filters(rows)
        
        return jsonify({
            'success': True,
            'message': f'File "{filename}" uploaded successfully! ({len(rows)} rows loaded)',
            'row_count': len(rows)
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'Upload failed: {str(e)}'}), 500


@app.route('/filters', methods=['GET'])
def get_filters():
    """Return computed filter options for the currently loaded file."""
    if not loaded_data.get('rows'):
        return jsonify({'success': False, 'message': 'No file loaded', 'filters': {}}), 400

    return jsonify({'success': True, 'filters': loaded_data.get('filters', {})})


def _apply_filters(rows, filters):
    """Filter rows by provided filter values. Filters are strings or empty."""
    if not filters:
        return rows
    # Filters can be single string or list of strings. When list: match any (OR) within that filter.
    mf = filters.get('manufacturer')
    pd = filters.get('product_division')
    ss = filters.get('sales_status')
    pm = filters.get('product_manager')
    si = filters.get('sub_item')
    mg = filters.get('material_group')
    mgd = filters.get('material_group_desc')

    def _matches_value(row_val, filter_val):
        if not filter_val:
            return True
        def _is_blank(v):
            return v is None or str(v).strip() == ''

        # If filter is a list, match any item in the list (OR). Support special '(blank)' token.
        if isinstance(filter_val, (list, tuple)):
            for f in filter_val:
                if f is None:
                    continue
                fstr = str(f).strip()
                if fstr == BLANK_LABEL and _is_blank(row_val):
                    return True
                if (row_val is not None) and str(row_val).strip() == fstr:
                    return True
            return False

        # Single filter value
        fstr = str(filter_val).strip()
        if fstr == BLANK_LABEL:
            return _is_blank(row_val)
        if row_val is None:
            return False
        return str(row_val).strip() == fstr

    def matches(row):
        # rows are tuples
        if not _matches_value(row[FIELD_IDX['manufacturer_name']] if len(row) > FIELD_IDX['manufacturer_name'] else None, mf):
            return False
        if not _matches_value(row[FIELD_IDX['product_division']] if len(row) > FIELD_IDX['product_division'] else None, pd):
            return False
        if not _matches_value(row[FIELD_IDX['sales_status']] if len(row) > FIELD_IDX['sales_status'] else None, ss):
            return False
        if not _matches_value(row[FIELD_IDX['material_group']] if len(row) > FIELD_IDX['material_group'] else None, mg):
            return False
        if not _matches_value(row[FIELD_IDX['material_group_desc']] if len(row) > FIELD_IDX['material_group_desc'] else None, mgd):
            return False
        if not _matches_value(row[FIELD_IDX['product_manager']] if len(row) > FIELD_IDX['product_manager'] else None, pm):
            return False
        if not _matches_value(row[FIELD_IDX['sub_item']] if len(row) > FIELD_IDX['sub_item'] else None, si):
            return False
        return True

    return [r for r in rows if matches(r)]

@app.route('/search', methods=['POST'])
def search():
    """Handle search request"""
    try:
        if not loaded_data['rows']:
            return jsonify({
                'success': False,
                'message': 'No file loaded. Please upload a file first.',
                'results': []
            }), 400
        
        data = request.get_json() or {}
        keywords = data.get('keywords', '').strip()
        filters = data.get('filters', {})

        # Apply filters first (if any)
        filtered_rows = _apply_filters(loaded_data['rows'], filters)

        # If keywords provided, perform keyword search on filtered rows
        if keywords:
            results = search_rows(keywords, filtered_rows)
        else:
            # No keywords => return filtered rows (or message prompting keywords if no filters)
            if filters:
                results = filtered_rows
            else:
                return jsonify({
                    'success': True,
                    'message': 'Please enter search keywords or apply filters',
                    'results': [],
                    'count': 0
                })
        
        if not results:
            return jsonify({
                'success': True,
                'message': 'No Match Found',
                'results': [],
                'count': 0,
                'no_match': True
            })
        
        # Convert matched tuple rows to dicts for response (do not convert entire dataset)
        def to_dict(tpl):
            return {k: (tpl[i] if i < len(tpl) else '') for i, k in enumerate(FIELDS)}

        result_dicts = [to_dict(r) for r in results]

        return jsonify({
            'success': True,
            'message': f'Found {len(result_dicts)} result(s)',
            'results': result_dicts,
            'count': len(result_dicts)
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Search failed: {str(e)}',
            'results': []
        }), 500


@app.route('/export', methods=['POST'])
def export_results():
    """Export search results (same logic as /search) to an Excel file and return as attachment."""
    try:
        if not loaded_data['rows']:
            return jsonify({'success': False, 'message': 'No file loaded. Please upload a file first.'}), 400

        data = request.get_json() or {}
        keywords = data.get('keywords', '').strip()
        filters = data.get('filters', {})

        # Apply filters
        filtered_rows = _apply_filters(loaded_data['rows'], filters)

        # Apply keyword search if provided
        if keywords:
            results = search_rows(keywords, filtered_rows)
        else:
            if filters:
                results = filtered_rows
            else:
                return jsonify({'success': False, 'message': 'Please enter search keywords or apply filters to export.'}), 400

        if not results:
            return jsonify({'success': True, 'message': 'No Match Found', 'results': [], 'count': 0}), 200

        # Create an Excel workbook in-memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Search Results'

        headers = ['Item No', 'Description', 'Product Division', 'Material Group', 'Material Group Desc', 'Manufacturer Name', 'Manufacturer Item No', 'Sales Status', 'Product Manager', 'Sub Item']
        ws.append(headers)

        # results are tuples; map by FIELD_IDX to the output column order
        for r in results:
            def gv(field):
                return (r[FIELD_IDX[field]] if len(r) > FIELD_IDX[field] else '')

            ws.append([
                gv('item_no'),
                gv('description'),
                gv('product_division'),
                gv('material_group'),
                gv('material_group_desc'),
                gv('manufacturer_name'),
                gv('manufacturer_item_no'),
                gv('sales_status'),
                gv('product_manager'),
                gv('sub_item')
            ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = 'search_results.xlsx'
        # Return file as attachment
        return send_file(bio,
                 as_attachment=True,
                 download_name=filename,
                 mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'success': False, 'message': f'Export failed: {str(e)}'}), 500

@app.route('/clear', methods=['POST'])
def clear():
    """Clear loaded file and search results"""
    loaded_data['rows'] = []
    loaded_data['filename'] = None
    return jsonify({'success': True, 'message': 'Data cleared. Ready for new upload.'})

if __name__ == '__main__':
    # Create uploads folder if it doesn't exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)
